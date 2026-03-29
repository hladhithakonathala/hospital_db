from flask import Flask, render_template, redirect, url_for, request, flash, session, send_file, jsonify
from flask_socketio import SocketIO, emit, join_room, leave_room
from flask_sqlalchemy import SQLAlchemy
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

app = Flask(__name__)
app.secret_key = 'apollo_hospital_secret_2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///hospital.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

app.config['MAIL_SERVER']         = 'smtp.gmail.com'
app.config['MAIL_PORT']           = 587
app.config['MAIL_USE_TLS']        = True
app.config['MAIL_USERNAME']       = 'your_email@gmail.com'
app.config['MAIL_PASSWORD']       = 'your_app_password'
app.config['MAIL_DEFAULT_SENDER'] = 'your_email@gmail.com'

mail     = Mail(app)
db       = SQLAlchemy(app)
socketio = SocketIO(app)

# ── Models ──────────────────────────────────────────────

class Patient(db.Model):
    id       = db.Column(db.Integer, primary_key=True)
    name     = db.Column(db.String(100), nullable=False)
    email    = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    age      = db.Column(db.Integer)
    phone    = db.Column(db.String(15))

class Doctor(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    name           = db.Column(db.String(100), nullable=False)
    email          = db.Column(db.String(120), unique=True, nullable=False)
    password       = db.Column(db.String(200), nullable=False)
    specialization = db.Column(db.String(100))

class Admin(db.Model):
    id       = db.Column(db.Integer, primary_key=True)
    name     = db.Column(db.String(100), nullable=False)
    email    = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)

class MedicalHistory(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    patient_id = db.Column(db.Integer, db.ForeignKey('patient.id'))
    date       = db.Column(db.String(50), nullable=False)
    diagnosis  = db.Column(db.String(200), nullable=False)
    treatment  = db.Column(db.String(200), nullable=False)
    doctor     = db.Column(db.String(100), nullable=False)
    notes      = db.Column(db.String(300))
    patient    = db.relationship('Patient', backref='medical_history')

class AvailableDate(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    doctor_id = db.Column(db.Integer, db.ForeignKey('doctor.id'))
    date      = db.Column(db.String(50), nullable=False)
    doctor    = db.relationship('Doctor', backref='available_dates')

class Offer(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    title       = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(200), nullable=False)
    discount    = db.Column(db.String(50), nullable=False)
    icon        = db.Column(db.String(10), default='🏥')

class Consultation(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    patient_id = db.Column(db.Integer, db.ForeignKey('patient.id'))
    doctor_id  = db.Column(db.Integer, db.ForeignKey('doctor.id'))
    date       = db.Column(db.String(50), nullable=False)
    time       = db.Column(db.String(50), nullable=False)
    reason     = db.Column(db.String(200), nullable=False)
    status     = db.Column(db.String(20), default='Pending')
    meet_link  = db.Column(db.String(200), default='')
    patient    = db.relationship('Patient', backref='consultations')
    doctor     = db.relationship('Doctor',  backref='consultations')

class ChatMessage(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    room        = db.Column(db.String(100), nullable=False)
    sender_id   = db.Column(db.Integer, nullable=False)
    sender_name = db.Column(db.String(100), nullable=False)
    sender_role = db.Column(db.String(20), nullable=False)
    message     = db.Column(db.Text, nullable=False)
    timestamp   = db.Column(db.String(50), nullable=False)
class Feedback(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    patient_id = db.Column(db.Integer, db.ForeignKey('patient.id'))
    doctor_id  = db.Column(db.Integer, db.ForeignKey('doctor.id'))
    rating     = db.Column(db.Integer, nullable=False)
    comment    = db.Column(db.String(500), nullable=False)
    date       = db.Column(db.String(50), nullable=False)
    patient    = db.relationship('Patient', backref='feedbacks')
    doctor     = db.relationship('Doctor',  backref='feedbacks')
class Appointment(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    patient_id = db.Column(db.Integer, db.ForeignKey('patient.id'))
    doctor_id  = db.Column(db.Integer, db.ForeignKey('doctor.id'))
    date       = db.Column(db.String(50))
    status     = db.Column(db.String(20), default='Pending')
    patient    = db.relationship('Patient', backref='appointments')
    doctor     = db.relationship('Doctor',  backref='appointments')

# ── Routes ──────────────────────────────────────────────

@app.route('/')
def index():
    offers = Offer.query.all()
    return render_template('index.html', offers=offers)

@app.route('/doctors')
def doctors_page():
    doctors = Doctor.query.all()
    return render_template('doctors.html', doctors=doctors)

@app.route('/login/<role>', methods=['GET', 'POST'])
def login(role):
    if request.method == 'POST':
        email    = request.form['email']
        password = request.form['password']
        if role == 'patient':
            user = Patient.query.filter_by(email=email).first()
        else:
            user = Doctor.query.filter_by(email=email).first()
        if user and check_password_hash(user.password, password):
            session['user_id']   = user.id
            session['user_role'] = role
            session['user_name'] = user.name
            return redirect(url_for(f'{role}_dashboard'))
        flash('Invalid email or password', 'error')
    return render_template('login.html', role=role)

@app.route('/register/<role>', methods=['GET', 'POST'])
def register(role):
    if request.method == 'POST':
        name     = request.form['name']
        email    = request.form['email']
        password = generate_password_hash(request.form['password'])
        if role == 'patient':
            user = Patient(name=name, email=email, password=password,
                           age=request.form.get('age'),
                           phone=request.form.get('phone'))
        else:
            user = Doctor(name=name, email=email, password=password,
                          specialization=request.form.get('specialization'))
        db.session.add(user)
        db.session.commit()
        flash('Registration successful! Please login.', 'success')
        return redirect(url_for('login', role=role))
    return render_template('login.html', role=role, register=True)

@app.route('/patient/dashboard')
def patient_dashboard():
    if session.get('user_role') != 'patient':
        return redirect(url_for('index'))
    patient      = Patient.query.get(session['user_id'])
    appointments = Appointment.query.filter_by(patient_id=patient.id).all()
    doctors      = Doctor.query.all()
    return render_template('patient_dashboard.html', patient=patient,
                           appointments=appointments, doctors=doctors)

@app.route('/doctor/dashboard')
def doctor_dashboard():
    if session.get('user_role') != 'doctor':
        return redirect(url_for('index'))
    doctor          = Doctor.query.get(session['user_id'])
    appointments    = Appointment.query.filter_by(doctor_id=doctor.id).all()
    available_dates = AvailableDate.query.filter_by(doctor_id=doctor.id).all()
    return render_template('doctor_dashboard.html', doctor=doctor,
                           appointments=appointments,
                           available_dates=available_dates)

@app.route('/book_appointment', methods=['POST'])
def book_appointment():
    appt = Appointment(
        patient_id=session['user_id'],
        doctor_id=request.form['doctor_id'],
        date=request.form['date']
    )
    db.session.add(appt)
    db.session.commit()
    flash('Appointment booked!', 'success')
    return redirect(url_for('patient_dashboard'))

@app.route('/update_appointment/<int:appt_id>/<action>')
def update_appointment(appt_id, action):
    if session.get('user_role') != 'doctor':
        return redirect(url_for('index'))
    appt = Appointment.query.get(appt_id)
    if action == 'confirm':
        appt.status = 'Confirmed'
    elif action == 'cancel':
        appt.status = 'Cancelled'
    db.session.commit()
    try:
        if action == 'confirm':
            subject = '✅ Appointment Confirmed - GVP Care'
            body = f'''Dear {appt.patient.name},

Your appointment has been CONFIRMED! 🎉

Details:
👨‍⚕️ Doctor: Dr. {appt.doctor.name}
🏥 Specialization: {appt.doctor.specialization}
📅 Date: {appt.date}
✅ Status: Confirmed

Please arrive 10 minutes early.

Thank you for choosing GVP Care!
🏥 GVP Care Team'''
        else:
            subject = '❌ Appointment Cancelled - GVP Care'
            body = f'''Dear {appt.patient.name},

Your appointment has been CANCELLED.

Details:
👨‍⚕️ Doctor: Dr. {appt.doctor.name}
📅 Date: {appt.date}
❌ Status: Cancelled

Please login to book a new appointment.

Thank you for choosing GVP Care!
🏥 GVP Care Team'''
        msg = Message(subject=subject,
                      recipients=[appt.patient.email],
                      body=body)
        mail.send(msg)
        flash(f'Appointment {action}ed & email sent!', 'success')
    except Exception as e:
        flash(f'Appointment {action}ed but email failed: {str(e)}', 'error')
    return redirect(url_for('doctor_dashboard'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    if not session.get('user_role'):
        return redirect(url_for('index'))
    if request.method == 'POST':
        current = request.form['current_password']
        new     = request.form['new_password']
        confirm = request.form['confirm_password']
        role    = session.get('user_role')
        if role == 'patient':
            user = Patient.query.get(session['user_id'])
        elif role == 'doctor':
            user = Doctor.query.get(session['user_id'])
        else:
            return redirect(url_for('index'))
        if not check_password_hash(user.password, current):
            flash('Current password is incorrect!', 'error')
        elif new != confirm:
            flash('New passwords do not match!', 'error')
        elif len(new) < 6:
            flash('New password must be at least 6 characters!', 'error')
        else:
            user.password = generate_password_hash(new)
            db.session.commit()
            flash('Password changed successfully!', 'success')
            return redirect(url_for(f'{role}_dashboard'))
    return render_template('change_password.html')

@app.route('/patient/profile', methods=['GET', 'POST'])
def patient_profile():
    if session.get('user_role') != 'patient':
        return redirect(url_for('index'))
    patient = Patient.query.get(session['user_id'])
    if request.method == 'POST':
        patient.name  = request.form['name']
        patient.age   = request.form['age']
        patient.phone = request.form['phone']
        db.session.commit()
        session['user_name'] = patient.name
        flash('Profile updated successfully!', 'success')
        return redirect(url_for('patient_profile'))
    return render_template('patient_profile.html', patient=patient)

@app.route('/patient/medical_history')
def medical_history():
    if session.get('user_role') != 'patient':
        return redirect(url_for('index'))
    patient = Patient.query.get(session['user_id'])
    history = MedicalHistory.query.filter_by(
        patient_id=patient.id).order_by(MedicalHistory.id.desc()).all()
    return render_template('medical_history.html',
                           patient=patient, history=history)

@app.route('/patient/add_medical_history', methods=['POST'])
def add_medical_history():
    if session.get('user_role') != 'patient':
        return redirect(url_for('index'))
    record = MedicalHistory(
        patient_id=session['user_id'],
        date=request.form['date'],
        diagnosis=request.form['diagnosis'],
        treatment=request.form['treatment'],
        doctor=request.form['doctor'],
        notes=request.form.get('notes', '')
    )
    db.session.add(record)
    db.session.commit()
    flash('Medical record added successfully!', 'success')
    return redirect(url_for('medical_history'))

@app.route('/patient/delete_medical_history/<int:id>')
def delete_medical_history(id):
    if session.get('user_role') != 'patient':
        return redirect(url_for('index'))
    record = MedicalHistory.query.get(id)
    if record.patient_id == session['user_id']:
        db.session.delete(record)
        db.session.commit()
        flash('Medical record deleted!', 'success')
    return redirect(url_for('medical_history'))

@app.route('/doctor/patient_history/<int:patient_id>')
def view_patient_history(patient_id):
    if session.get('user_role') != 'doctor':
        return redirect(url_for('index'))
    patient = Patient.query.get(patient_id)
    history = MedicalHistory.query.filter_by(
        patient_id=patient_id).order_by(MedicalHistory.id.desc()).all()
    return render_template('medical_history.html',
                           patient=patient, history=history, view_only=True)

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        email    = request.form['email']
        password = request.form['password']
        admin    = Admin.query.filter_by(email=email).first()
        if admin and check_password_hash(admin.password, password):
            session['user_id']   = admin.id
            session['user_role'] = 'admin'
            session['user_name'] = admin.name
            return redirect(url_for('admin_dashboard'))
        flash('Invalid email or password', 'error')
    return render_template('admin_login.html')

@app.route('/admin/register', methods=['GET', 'POST'])
def admin_register():
    if request.method == 'POST':
        name     = request.form['name']
        email    = request.form['email']
        password = generate_password_hash(request.form['password'])
        admin    = Admin(name=name, email=email, password=password)
        db.session.add(admin)
        db.session.commit()
        flash('Admin registered! Please login.', 'success')
        return redirect(url_for('admin_login'))
    return render_template('admin_login.html', register=True)

@app.route('/admin/dashboard')
def admin_dashboard():
    if session.get('user_role') != 'admin':
        return redirect(url_for('index'))
    patients     = Patient.query.all()
    doctors      = Doctor.query.all()
    appointments = Appointment.query.all()
    offers       = Offer.query.all()
    return render_template('admin_dashboard.html',
                           patients=patients,
                           doctors=doctors,
                           appointments=appointments,
                           offers=offers)

@app.route('/admin/delete_patient/<int:id>')
def delete_patient(id):
    if session.get('user_role') != 'admin':
        return redirect(url_for('index'))
    patient = Patient.query.get(id)
    Appointment.query.filter_by(patient_id=id).delete()
    MedicalHistory.query.filter_by(patient_id=id).delete()
    db.session.delete(patient)
    db.session.commit()
    flash('Patient deleted!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete_doctor/<int:id>')
def delete_doctor(id):
    if session.get('user_role') != 'admin':
        return redirect(url_for('index'))
    doctor = Doctor.query.get(id)
    Appointment.query.filter_by(doctor_id=id).delete()
    AvailableDate.query.filter_by(doctor_id=id).delete()
    db.session.delete(doctor)
    db.session.commit()
    flash('Doctor deleted!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/add_offer', methods=['POST'])
def add_offer():
    if session.get('user_role') != 'admin':
        return redirect(url_for('index'))
    offer = Offer(
        title=request.form['title'],
        description=request.form['description'],
        discount=request.form['discount'],
        icon=request.form['icon']
    )
    db.session.add(offer)
    db.session.commit()
    flash('Offer added successfully!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete_offer/<int:id>')
def delete_offer(id):
    if session.get('user_role') != 'admin':
        return redirect(url_for('index'))
    offer = Offer.query.get(id)
    db.session.delete(offer)
    db.session.commit()
    flash('Offer deleted!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/export')
def export_excel():
    if session.get('user_role') != 'admin':
        return redirect(url_for('index'))
    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Patients'
    ws1.append(['ID', 'Name', 'Email', 'Age', 'Phone'])
    for cell in ws1[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='003366')
        cell.alignment = Alignment(horizontal='center')
    for p in Patient.query.all():
        ws1.append([p.id, p.name, p.email, p.age or '—', p.phone or '—'])
    ws2 = wb.create_sheet('Doctors')
    ws2.append(['ID', 'Name', 'Email', 'Specialization'])
    for cell in ws2[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='e63946')
        cell.alignment = Alignment(horizontal='center')
    for d in Doctor.query.all():
        ws2.append([d.id, d.name, d.email, d.specialization or '—'])
    ws3 = wb.create_sheet('Appointments')
    ws3.append(['ID', 'Patient', 'Doctor', 'Date', 'Status'])
    for cell in ws3[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='2d7a4f')
        cell.alignment = Alignment(horizontal='center')
    for a in Appointment.query.all():
        ws3.append([a.id, a.patient.name, a.doctor.name, a.date, a.status])
    ws4 = wb.create_sheet('Medical History')
    ws4.append(['ID', 'Patient', 'Date', 'Diagnosis', 'Treatment', 'Doctor', 'Notes'])
    for cell in ws4[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='b06000')
        cell.alignment = Alignment(horizontal='center')
    for h in MedicalHistory.query.all():
        ws4.append([h.id, h.patient.name, h.date,
                    h.diagnosis, h.treatment, h.doctor, h.notes or '—'])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     download_name='GVPCare_Data.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/search')
def admin_search():
    if session.get('user_role') != 'admin':
        return redirect(url_for('index'))
    query    = request.args.get('q', '')
    patients = Patient.query.filter(Patient.name.ilike(f'%{query}%')).all()
    doctors  = Doctor.query.filter(Doctor.name.ilike(f'%{query}%')).all()
    return render_template('admin_search.html',
                           patients=patients,
                           doctors=doctors,
                           query=query)

@app.route('/admin/add_doctor', methods=['POST'])
def admin_add_doctor():
    if session.get('user_role') != 'admin':
        return redirect(url_for('index'))
    name           = request.form['name']
    email          = request.form['email']
    password       = generate_password_hash(request.form['password'])
    specialization = request.form['specialization']
    doctor = Doctor(name=name, email=email,
                    password=password,
                    specialization=specialization)
    db.session.add(doctor)
    db.session.commit()
    flash('Doctor added successfully!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/update_appointment/<int:appt_id>/<status>')
def admin_update_appointment(appt_id, status):
    if session.get('user_role') != 'admin':
        return redirect(url_for('index'))
    appt        = Appointment.query.get(appt_id)
    appt.status = status
    db.session.commit()
    flash(f'Appointment status updated to {status}!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete_appointment/<int:id>')
def delete_appointment(id):
    if session.get('user_role') != 'admin':
        return redirect(url_for('index'))
    appt = Appointment.query.get(id)
    db.session.delete(appt)
    db.session.commit()
    flash('Appointment deleted!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/doctor/add_available_date', methods=['POST'])
def add_available_date():
    if session.get('user_role') != 'doctor':
        return redirect(url_for('index'))
    date = AvailableDate(
        doctor_id=session['user_id'],
        date=request.form['date']
    )
    db.session.add(date)
    db.session.commit()
    flash('Available date added!', 'success')
    return redirect(url_for('doctor_dashboard'))

@app.route('/doctor/delete_available_date/<int:id>')
def delete_available_date(id):
    if session.get('user_role') != 'doctor':
        return redirect(url_for('index'))
    date = AvailableDate.query.get(id)
    db.session.delete(date)
    db.session.commit()
    flash('Date removed!', 'success')
    return redirect(url_for('doctor_dashboard'))

@app.route('/api/available_dates/<int:doctor_id>')
def get_available_dates(doctor_id):
    dates = AvailableDate.query.filter_by(doctor_id=doctor_id).all()
    return jsonify([d.date for d in dates])

@app.route('/api/appointments')
def get_appointments():
    if session.get('user_role') == 'patient':
        appts = Appointment.query.filter_by(patient_id=session['user_id']).all()
    elif session.get('user_role') == 'doctor':
        appts = Appointment.query.filter_by(doctor_id=session['user_id']).all()
    else:
        appts = Appointment.query.all()
    return jsonify([{
        'title': f"{a.patient.name} → Dr. {a.doctor.name}",
        'start': a.date,
        'color': '#2d7a4f' if a.status == 'Confirmed' else '#e63946' if a.status == 'Cancelled' else '#b06000'
    } for a in appts])

@app.route('/chat/<int:doctor_id>')
def chat(doctor_id):
    if session.get('user_role') not in ['patient', 'doctor']:
        return redirect(url_for('index'))
    doctor     = Doctor.query.get(doctor_id)
    patient_id = request.args.get('patient_id', type=int)
    if session.get('user_role') == 'doctor':
        patient = Patient.query.get(patient_id)
    else:
        patient = Patient.query.get(session['user_id'])
    room     = f"chat_{min(doctor_id, patient.id)}_{max(doctor_id, patient.id)}"
    messages = ChatMessage.query.filter_by(room=room).all()
    return render_template('chat.html', doctor=doctor, patient=patient,
                           room=room, messages=messages)

@app.route('/consultation/request', methods=['GET', 'POST'])
def request_consultation():
    if session.get('user_role') != 'patient':
        return redirect(url_for('index'))
    doctors = Doctor.query.all()
    if request.method == 'POST':
        consultation = Consultation(
            patient_id=session['user_id'],
            doctor_id=request.form['doctor_id'],
            date=request.form['date'],
            time=request.form['time'],
            reason=request.form['reason']
        )
        db.session.add(consultation)
        db.session.commit()
        flash('Online consultation requested successfully!', 'success')
        return redirect(url_for('my_consultations'))
    return render_template('request_consultation.html', doctors=doctors)

@app.route('/consultation/my')
def my_consultations():
    if session.get('user_role') != 'patient':
        return redirect(url_for('index'))
    patient       = Patient.query.get(session['user_id'])
    consultations = Consultation.query.filter_by(
        patient_id=session['user_id']).order_by(Consultation.id.desc()).all()
    return render_template('my_consultations.html',
                           patient=patient,
                           consultations=consultations)

@app.route('/consultation/doctor')
def doctor_consultations():
    if session.get('user_role') != 'doctor':
        return redirect(url_for('index'))
    doctor        = Doctor.query.get(session['user_id'])
    consultations = Consultation.query.filter_by(
        doctor_id=session['user_id']).order_by(Consultation.id.desc()).all()
    return render_template('doctor_consultations.html',
                           doctor=doctor,
                           consultations=consultations)

@app.route('/consultation/update/<int:id>', methods=['POST'])
def update_consultation(id):
    if session.get('user_role') != 'doctor':
        return redirect(url_for('index'))
    consultation           = Consultation.query.get(id)
    consultation.status    = request.form['status']
    consultation.meet_link = request.form.get('meet_link', '')
    db.session.commit()
    flash('Consultation updated successfully!', 'success')
    return redirect(url_for('doctor_consultations'))

@app.route('/consultation/cancel/<int:id>')
def cancel_consultation(id):
    if session.get('user_role') != 'patient':
        return redirect(url_for('index'))
    consultation = Consultation.query.get(id)
    if consultation.patient_id == session['user_id']:
        consultation.status = 'Cancelled'
        db.session.commit()
        flash('Consultation cancelled!', 'success')
    return redirect(url_for('my_consultations'))

# ── SocketIO Events ──────────────────────────────────────

@socketio.on('join')
def on_join(data):
    join_room(data['room'])

@socketio.on('leave')
def on_leave(data):
    leave_room(data['room'])

@socketio.on('send_message')
def handle_message(data):
    from datetime import datetime
    timestamp = datetime.now().strftime('%I:%M %p')
    msg = ChatMessage(
        room=data['room'],
        sender_id=data['sender_id'],
        sender_name=data['sender_name'],
        sender_role=data['sender_role'],
        message=data['message'],
        timestamp=timestamp
    )
    db.session.add(msg)
    db.session.commit()
    emit('receive_message', {
        'sender_name': data['sender_name'],
        'sender_role': data['sender_role'],
        'message':     data['message'],
        'timestamp':   timestamp
    }, room=data['room'])
@app.route('/patient/feedback', methods=['GET', 'POST'])
def patient_feedback():
    if session.get('user_role') != 'patient':
        return redirect(url_for('index'))
    patient  = Patient.query.get(session['user_id'])
    doctors  = Doctor.query.all()
    feedbacks = Feedback.query.filter_by(
        patient_id=patient.id).order_by(Feedback.id.desc()).all()
    if request.method == 'POST':
        # Check if already reviewed this doctor
        existing = Feedback.query.filter_by(
            patient_id=patient.id,
            doctor_id=request.form['doctor_id']).first()
        if existing:
            flash('You have already reviewed this doctor!', 'error')
        else:
            from datetime import date
            feedback = Feedback(
                patient_id=patient.id,
                doctor_id=request.form['doctor_id'],
                rating=int(request.form['rating']),
                comment=request.form['comment'],
                date=str(date.today())
            )
            db.session.add(feedback)
            db.session.commit()
            flash('Thank you for your feedback! ⭐', 'success')
        return redirect(url_for('patient_feedback'))
    return render_template('patient_feedback.html',
                           patient=patient,
                           doctors=doctors,
                           feedbacks=feedbacks)

@app.route('/patient/delete_feedback/<int:id>')
def delete_feedback(id):
    if session.get('user_role') != 'patient':
        return redirect(url_for('index'))
    feedback = Feedback.query.get(id)
    if feedback.patient_id == session['user_id']:
        db.session.delete(feedback)
        db.session.commit()
        flash('Feedback deleted!', 'success')
    return redirect(url_for('patient_feedback'))

@app.route('/doctor/feedbacks')
def doctor_feedbacks():
    if session.get('user_role') != 'doctor':
        return redirect(url_for('index'))
    doctor    = Doctor.query.get(session['user_id'])
    feedbacks = Feedback.query.filter_by(
        doctor_id=doctor.id).order_by(Feedback.id.desc()).all()
    avg_rating = round(sum(f.rating for f in feedbacks) / len(feedbacks), 1) if feedbacks else 0
    return render_template('doctor_feedbacks.html',
                           doctor=doctor,
                           feedbacks=feedbacks,
                           avg_rating=avg_rating)
# ── Init DB & Run ────────────────────────────────────────

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    socketio.run(app, debug=True, host='0.0.0.0', allow_unsafe_werkzeug=True) 